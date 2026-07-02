from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "Issues Register.xlsm"
OUTPUT_PATH = ROOT / "issues-data.js"


FIELD_MAP = {
    "ID (Unique Identifier)": "id",
    "Use Cases": "useCasesImpacted",
    "Category": "category",
    "Journey Stage": "journeyStage",
    "Issue Title": "title",
    "Problem Scenario\n[context] + [problem] + [impact] + [current state] + [Direction]": "problemScenario",
    "Issue Type": "issueType",
    "Dependencies\n": "dependencies",
    "Related Initiative / PE / Jira / Confluence": "relatedInitiative",
    "Product": "productsImpacted",
    "Product Owner (Accountable)": "owner",
    "Priority to unlock the use case": "priority",
    "Time Horizon Short Term - 0 -6 months\nMedium Term - 6 months -2 years\nLong Term - 2 years+": "horizon",
    "Status": "status",
    "Current Action": "currentAction",
    "Decision Needed": "decisionNeeded",
    "Comments / Notes": "notes",
    "Link 1\n(details of the issue)": "link1",
    "Link 2": "link2",
    "Strategic Priority ID": "strategicPriorityId",
    "Objective ID": "objectiveId",
    "Scorecard ID": "scorecardId",
    "Last Updated": "lastUpdated",
    "Exec Summary Tag": "execSummaryTag",
}

OPTIONAL_FIELD_MAP = {
    "Strategic Objective": "strategicObjective",
    "Scorecard Item": "scorecardItem",
    "Readiness Status": "readinessStatus",
    "Readiness Rationale": "readinessRationale",
    "Decision Owner": "decisionOwner",
    "Decision Due Date": "decisionDueDate",
    "Dependency Owner Type": "dependencyOwnerType",
    "Impact Driver": "impactDriver",
    "Roadmap Alignment": "roadmapAlignment",
    "External Messaging Status": "externalMessagingStatus",
    "Review Date": "reviewDate",
}

USE_CASE_ALIASES = {
    "all": "All",
    "super": "Super",
    "payday super": "Super",
    "payroll": "Payroll",
    "tax": "Tax",
    "econveyancing": "eConveyancing",
    "electronic conveyancing": "eConveyancing",
    "electronic conveyancing econveyancing": "eConveyancing",
    "einvoicing": "eInvoicing",
    "e invoicing": "eInvoicing",
    "securities": "Securities",
    "government": "Government",
    "business payments": "Business Payments",
    "bp": "Business Payments",
    "mddr": "MDDR",
}


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize(value):
    text = clean_text(value)
    return text.upper() if text else ""


def split_values(*values):
    items = []
    seen = set()
    for value in values:
        text = "" if value is None else str(value).strip()
        if not text:
            continue
        for item in re.split(r"\s*(?:;|\n)\s*", text):
            item = clean_text(item)
            key = item.lower()
            if item and key not in seen:
                items.append(item)
                seen.add(key)
    return items


def normalize_use_case(value):
    text = clean_text(value)
    key = re.sub(r"[^a-z0-9]+", " ", text.lower().replace("&", "and")).strip()
    return USE_CASE_ALIASES.get(key, text)


def normalize_key(value):
    return clean_text(value).lower()


def read_sheet_records(workbook, sheet_name, field_map):
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        if not any(value is not None for value in raw.values()):
            continue

        record = {}
        for source, target in field_map.items():
            sources = source if isinstance(source, tuple) else (source,)
            record[target] = ""
            for candidate in sources:
                value = clean_text(raw.get(candidate))
                if value:
                    record[target] = value
                    break
        if any(record.values()):
            records.append(record)
    return records


def build_lookup(records, field):
    return {normalize_key(record.get(field)): record for record in records if record.get(field)}


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["IssuesRegister"] if "IssuesRegister" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in sheet[1]]

    strategy = read_sheet_records(
        workbook,
        "Ref_Strategy",
        {
            "Strategic Priority ID": "id",
            "Strategic Priority": "title",
        },
    )
    objectives = read_sheet_records(
        workbook,
        "Ref_Objectives",
        {
            "Objective ID": "id",
            "Objective": "title",
            "Description": "description",
        },
    )
    scorecard = read_sheet_records(
        workbook,
        "Ref_Scorecard",
        {
            "Scorecard ID": "id",
            "Key Result": "keyResult",
            "Threshold": "threshold",
            "Target": "target",
            "Maximum": "maximum",
            ("Use Case", "Use Cases", "Usa Case"): "useCasesImpacted",
        },
    )
    use_cases = read_sheet_records(
        workbook,
        "Ref_Use_Cases",
        {
            "Use Case": "name",
            "Description": "description",
            "Desired Outcome": "desiredOutcome",
            "Status": "status",
        },
    )
    capabilities = read_sheet_records(
        workbook,
        "Ref_Capabilities",
        {
            "Capability": "name",
            "Owner": "owner",
            "Description": "description",
        },
    )

    strategy_by_id = build_lookup(strategy, "id")
    objectives_by_id = build_lookup(objectives, "id")
    scorecard_by_id = build_lookup(scorecard, "id")
    use_cases_by_name = {normalize_key(normalize_use_case(item.get("name"))): item for item in use_cases if item.get("name")}
    capabilities_by_name = build_lookup(capabilities, "name")

    for item in scorecard:
        item["useCases"] = [normalize_use_case(value) for value in split_values(item.get("useCasesImpacted"))]

    issues = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        if not any(value is not None for value in raw.values()):
            continue

        issue = {}
        for source, target in FIELD_MAP.items():
            issue[target] = clean_text(raw.get(source))
        for source, target in OPTIONAL_FIELD_MAP.items():
            issue[target] = clean_text(raw.get(source))

        issue["priority"] = normalize(issue["priority"])
        issue["horizon"] = normalize(issue["horizon"])
        issue["status"] = normalize(issue["status"])
        issue["isExecSummary"] = issue["execSummaryTag"].lower() == "yes"

        issue["useCases"] = [normalize_use_case(value) for value in split_values(raw.get("Use Cases"))]
        issue["useCasesImpacted"] = "; ".join(issue["useCases"])
        issue["useCase"] = issue["useCasesImpacted"]
        issue["categories"] = split_values(raw.get("Category"))
        issue["category"] = "; ".join(issue["categories"])
        issue["products"] = split_values(raw.get("Product"))
        issue["productsImpacted"] = "; ".join(issue["products"])
        issue["product"] = issue["productsImpacted"]
        issue["journeyStages"] = split_values(raw.get("Journey Stage"))
        issue["journeyStage"] = "; ".join(issue["journeyStages"])
        issue["scorecardIds"] = split_values(raw.get("Scorecard ID"))
        issue["scorecardId"] = "; ".join(issue["scorecardIds"])

        issue["strategicPriority"] = strategy_by_id.get(normalize_key(issue.get("strategicPriorityId")), {})
        issue["objective"] = objectives_by_id.get(normalize_key(issue.get("objectiveId")), {})
        issue["scorecards"] = [
            scorecard_by_id.get(normalize_key(scorecard_id), {})
            for scorecard_id in issue["scorecardIds"]
            if scorecard_by_id.get(normalize_key(scorecard_id))
        ]
        issue["scorecard"] = issue["scorecards"][0] if issue["scorecards"] else {}
        issue["useCaseRefs"] = [
            use_cases_by_name.get(normalize_key(use_case), {})
            for use_case in issue["useCases"]
            if use_cases_by_name.get(normalize_key(use_case))
        ]
        issue["capabilityRefs"] = [
            capabilities_by_name.get(normalize_key(product), {})
            for product in issue["products"]
            if capabilities_by_name.get(normalize_key(product))
        ]
        issues.append(issue)

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": WORKBOOK_PATH.name,
        "strategy": strategy,
        "objectives": objectives,
        "scorecard": scorecard,
        "useCases": use_cases,
        "capabilities": capabilities,
        "issues": issues,
    }

    OUTPUT_PATH.write_text(
        "window.ISSUES_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(issues)} issues to {OUTPUT_PATH.name}")


if __name__ == "__main__":
    main()
